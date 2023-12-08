import datetime
import requests
import pandas as pd
import pprint
import time
import random
from bs4 import BeautifulSoup
import json
import csv
import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication, QTreeView, QFileSystemModel, QVBoxLayout, QPushButton, QInputDialog, \
    QLineEdit, QMainWindow, QMessageBox, QFileDialog
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime, date, timedelta
import numpy
import datetime
from window import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import json
import pprint
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.styles import Alignment,PatternFill
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity


def GetKeywords(fname):
    # 엑셀 파일 읽기

    df = pd.read_excel(fname)

    # 첫 번째 열만 선택하여 딕셔너리로 변환
    first_column = df.iloc[:, 0].str.strip()
    keywords = first_column.tolist()
    print("keywords:",keywords,"/ keywords_TYPE:",type(keywords),len(keywords))
    searchKeywords=" | ".join(keywords)
    print("searchKeywords:",searchKeywords,"/ searchKeywords_TYPE:",type(searchKeywords),len(searchKeywords))
    return keywords,searchKeywords



def GetCompanyName(fname):
    # 엑셀 파일 읽기

    df = pd.read_excel(fname)

    # 첫 번째 열만 선택하여 딕셔너리로 변환
    first_column = df.iloc[:, 0].str.strip()
    companyNames = first_column.tolist()

    pprint.pprint(companyNames)
    return companyNames

def remove_duplicate_dicts(lst, key):
    seen = set()
    result = []
    for d in lst:
        if d[key] not in seen:
            seen.add(d[key])
            result.append(d)
    return result

def GetArticles(keywords,searchKeywords,companyNames,searchStart,searchEnd,countLimit,self):
    count=0
    resultList=[]
    while True:
        text = "{}/{}번째 페이지 기사 가져오기".format(count+1,countLimit)
        self.user_signal.emit(text)
        cookies = {
            # 'NNB': 'EOQEJJQUFMYWK',
            # 'nx_ssl': '2',
            # 'ASID': '3d6968c30000018b481054af00008e99',
            # '_ga': 'GA1.2.1867842095.1698124262',
            # '_fbp': 'fb.1.1698124262166.1924394679',
            # '_tt_enable_cookie': '1',
            # '_ttp': 'BEmRKkFpqC2qDqQF7ScXQ2-Dz-t',
            # '_ga_4BKHBFKFK0': 'GS1.1.1698124262.1.1.1698124263.59.0.0',
            # 'nid_inf': '-1418340302',
            # 'NID_AUT': '3SLnksOBVKphncVoHj0APDbP5U/ED2YpQYleT1yokMIn+HQL9NuWvfzlBs8Vg6FE',
            # 'NID_JKL': 'UWjHrUZIdvQ8BXw7yulMx+BHWLqeT5p9oA1MkpzzkAw=',
            # '_naver_usersession_': 'UQOaZKGIfz29lcXjqQmiSA==',
            # 'NID_SES': 'AAABpCj5RNSPm+yCjzkae5bDyEbeQ5NrJ1dEqogMrSDUDKoyonJClHDOWMCc3KZL/ANrcNVFE+tRQm/JmswWWNzMPEcBZdWtRzlyrQK8bNWTtco1LekPfTTR6VBlFJxBw/4596ANGrhHafooOoqEHysuhiBOmjP3gDv5iXzDkUICCa4jGKIATaYtyJlzUCHPjFFuiTQHdfzADdRZdI1JiYPvugtdyJbAiXnFGkIjxxKlRaCI1DCL1zbgoEkrYmdzssUKmFmgxd+NQkqc4L9aolVkLvGmetw/y/9Y1hKhap3OWBbHzyYybX7uoaX1OExMcyCK8of+OEipk3/28r/0aym9WLpuiXIq3L7DFB+o5V3qK8JtXsPqqS8WDZ4JoHenw6zVPwVSvxXHqqLR9ThJ0amnUptiNmweairnpOoltCL1GCB+IVdAmTfzTDgGgaOBNROqmqxtW0RYUyPJQNaS0RumOQkkikKHGnp3brNIL71HEJwGOSXrZe1ZEtBt6APGzjqROCPYjSJNK2yACzkFyCBcrbkc9FyV4AbWXFexOw5l/0WeFZGdPtEUNcS2JceYx3IO9A==',
            # 'page_uid': 'ig0m5dp0J1sss4rZRedssssst3N-311207',
        }

        headers = {
            'authority': 'search.naver.com',
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'cache-control': 'max-age=0',
            # 'cookie': 'NNB=EOQEJJQUFMYWK; nx_ssl=2; ASID=3d6968c30000018b481054af00008e99; _ga=GA1.2.1867842095.1698124262; _fbp=fb.1.1698124262166.1924394679; _tt_enable_cookie=1; _ttp=BEmRKkFpqC2qDqQF7ScXQ2-Dz-t; _ga_4BKHBFKFK0=GS1.1.1698124262.1.1.1698124263.59.0.0; nid_inf=-1418340302; NID_AUT=3SLnksOBVKphncVoHj0APDbP5U/ED2YpQYleT1yokMIn+HQL9NuWvfzlBs8Vg6FE; NID_JKL=UWjHrUZIdvQ8BXw7yulMx+BHWLqeT5p9oA1MkpzzkAw=; _naver_usersession_=UQOaZKGIfz29lcXjqQmiSA==; NID_SES=AAABpCj5RNSPm+yCjzkae5bDyEbeQ5NrJ1dEqogMrSDUDKoyonJClHDOWMCc3KZL/ANrcNVFE+tRQm/JmswWWNzMPEcBZdWtRzlyrQK8bNWTtco1LekPfTTR6VBlFJxBw/4596ANGrhHafooOoqEHysuhiBOmjP3gDv5iXzDkUICCa4jGKIATaYtyJlzUCHPjFFuiTQHdfzADdRZdI1JiYPvugtdyJbAiXnFGkIjxxKlRaCI1DCL1zbgoEkrYmdzssUKmFmgxd+NQkqc4L9aolVkLvGmetw/y/9Y1hKhap3OWBbHzyYybX7uoaX1OExMcyCK8of+OEipk3/28r/0aym9WLpuiXIq3L7DFB+o5V3qK8JtXsPqqS8WDZ4JoHenw6zVPwVSvxXHqqLR9ThJ0amnUptiNmweairnpOoltCL1GCB+IVdAmTfzTDgGgaOBNROqmqxtW0RYUyPJQNaS0RumOQkkikKHGnp3brNIL71HEJwGOSXrZe1ZEtBt6APGzjqROCPYjSJNK2yACzkFyCBcrbkc9FyV4AbWXFexOw5l/0WeFZGdPtEUNcS2JceYx3IO9A==; page_uid=ig0m5dp0J1sss4rZRedssssst3N-311207',
            'referer': 'https://search.naver.com/search.naver?where=news&sm=tab_pge&query=%EC%82%BC%EC%84%B1%EC%A0%84%EC%9E%90&sort=1&photo=0&field=0&pd=12&ds=2023.10.28.17.56&de=2023.10.28.23.56&mynews=0&office_type=0&office_section_code=0&news_office_checked=&office_category=0&service_area=0&nso=so:dd,p:all,a:all&start=11',
            'sec-ch-ua': '"Chromium";v="118", "Google Chrome";v="118", "Not=A?Brand";v="99"',
            'sec-ch-ua-arch': '"x86"',
            'sec-ch-ua-bitness': '"64"',
            'sec-ch-ua-full-version-list': '"Chromium";v="118.0.5993.117", "Google Chrome";v="118.0.5993.117", "Not=A?Brand";v="99.0.0.0"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-model': '""',
            'sec-ch-ua-platform': '"Windows"',
            'sec-ch-ua-platform-version': '"10.0.0"',
            'sec-ch-ua-wow64': '?0',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'same-origin',
            'sec-fetch-user': '?1',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36',
        }

        params = {
            'where': 'news',
            'query': searchKeywords,
            'sm': 'tab_opt',
            'sort': '0',
            'photo': '0',
            'field': '0',
            'pd': '3',
            'ds': searchStart,
            'de': searchEnd,
            'docid': '',
            'related': '0',
            'mynews': '0',
            'office_type': '0',
            'office_section_code': '0',
            'news_office_checked': '',
            'nso': 'so:dd,p:from{}to{}'.format(searchStart.replace(".",""),searchEnd.replace(".","")),
            'is_sug_officeid': '0',
            'office_category': '0',
            'service_area': '0',
            'start':count*10+1
        }
        print("params:",params,"/ params_TYPE:",type(params),len(params))

        response = requests.get('https://search.naver.com/search.naver', params=params, cookies=cookies,
                                headers=headers)
        # print(response.text)
        soup=BeautifulSoup(response.text,'lxml')
        # print(soup.prettify())
        try:
            articleGroup=soup.find("ul",attrs={'class':'list_news'})
            articles = articleGroup.find_all('li', attrs={'class': 'bx'})
        except:
            print("기사없음1")
            break

        if len(articles)==0:
            print("더없음")
            break

        for article in articles:
            company=article.find('a',attrs={'class':'info press'}).get_text().replace("언론사 선정","")
            # print("company:",company,"/ company_TYPE:",type(company),len(company))
            title=article.find('a',attrs={'class':'news_tit'}).get_text()
            titleList=title.replace(",","").split(" ")
            allLinks=article.find_all('a')
            companyList=[]
            keywordFind=[]
            companyFind=[]
            findFlag=False
            pprint.pprint(titleList)
            for allLink in allLinks:
                if allLink['href'].find("n.news.naver.com")>=0:
                    for companyName in companyNames:
                        if companyName in titleList:
                            companyFind.append(companyName)
                            findFlag=True
                        
                        if findFlag==True:
                            # print("네이버뉴스/주요기업에해당함:{}".format(title))
                            pickInfo=""
                            url=allLink['href']
                            # print("url:",url,"/ url_TYPE:",type(url))
                            isPapers=article.find_all('span',attrs={'class':'info'})
                            if len(isPapers)>=1:
                                for isPaper in isPapers:
                                    if isPaper.get_text().find("면")>=0:
                                        pickInfo=pickInfo+" "+isPaper.get_text()

                            data={'title':title,'url':url,'company':company,'keyword':searchKeywords,'companyFind':companyFind}
                            # print("data:",data,"/ data_TYPE:",type(data),len(data))
                            resultList.append(data)
                    break
            # with open('resultList.json', 'w',encoding='utf-8-sig') as f:
            #     json.dump(resultList, f, indent=2,ensure_ascii=False)

        print("================{}/{}=============".format(count+1,countLimit))
        if count>=countLimit:
            break
        count+=1

        time.sleep(random.randint(10,20)*0.1)
    return resultList

def GetDetail(keywords,result):
    cookies = {
        'NNB': 'EOQEJJQUFMYWK',
        'nx_ssl': '2',
        'ASID': '3d6968c30000018b481054af00008e99',
        '_ga': 'GA1.2.1867842095.1698124262',
        '_fbp': 'fb.1.1698124262166.1924394679',
        '_tt_enable_cookie': '1',
        '_ttp': 'BEmRKkFpqC2qDqQF7ScXQ2-Dz-t',
        '_ga_4BKHBFKFK0': 'GS1.1.1698124262.1.1.1698124263.59.0.0',
        'page_uid': 'iSuIKdqo15VssE5ybkwssssstIV-153411',
        '_naver_usersession_': '+fHq3YE7rKngzMKSrj2eMQ==',
        'N_SES': '4fbe6a0a-f4dd-441e-8ec4-4d0074be9a9a',
        'VISIT_LOG_CLEAN': '1',
    }

    headers = {
        'authority': 'n.news.naver.com',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'cache-control': 'max-age=0',
        # 'cookie': 'NNB=EOQEJJQUFMYWK; nx_ssl=2; ASID=3d6968c30000018b481054af00008e99; _ga=GA1.2.1867842095.1698124262; _fbp=fb.1.1698124262166.1924394679; _tt_enable_cookie=1; _ttp=BEmRKkFpqC2qDqQF7ScXQ2-Dz-t; _ga_4BKHBFKFK0=GS1.1.1698124262.1.1.1698124263.59.0.0; page_uid=iSuIKdqo15VssE5ybkwssssstIV-153411; _naver_usersession_=+fHq3YE7rKngzMKSrj2eMQ==; N_SES=4fbe6a0a-f4dd-441e-8ec4-4d0074be9a9a; VISIT_LOG_CLEAN=1',
        'referer': 'https://search.naver.com/search.naver?where=news&ie=utf8&sm=nws_hty&query=%EC%A6%9D%EC%84%A4',
        'sec-ch-ua': '"Google Chrome";v="119", "Chromium";v="119", "Not?A_Brand";v="24"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'same-site',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
    }

    # params = {
    #     'sid': '102',
    # }

    response = requests.get(result['url'], cookies=cookies,
                            headers=headers)

    
    soup=BeautifulSoup(response.text,'lxml')
    # print(soup.prettify())
    if len(soup.find_all("div",attrs={'id':'newsct_article'}))>=1:
        print("경제")
        contents=soup.find("div",attrs={'id':'newsct_article'}).get_text().strip().replace("\n","")
        try:
            regiDate = soup.find("span",
                                 attrs={'class': 'media_end_head_info_datestamp_time _ARTICLE_DATE_TIME'}).get_text()
        except:
            regiDate = ""
        print("regiDate:", regiDate)

    elif len(soup.find_all("div",attrs={'id':'articeBody'}))>=1:

        contents = soup.find("div", attrs={'id': 'articeBody'}).get_text().strip().replace("\n","")

        try:
            regiDate=soup.find("div",attrs={'class':'article_info'}).find('em').get_text()
        except:
            regiDate=""
        print("regiDate:",regiDate)
    else:
        print("기타")
        contents=""
    print("contents:",contents,"/ contents_TYPE:",type(contents),len(contents))

    searchList=[]

    # 원본 문자열

    # 특수 문자를 기준으로 문자열 분할
    keywordsList = re.split(r'[^가-힣A-Za-z0-9]+', keywords)
    print("keywordsList:",keywordsList,"/ keywordsList_TYPE:",type(keywordsList))


    for keyword in keywordsList:
        if contents.find(keyword)>=0:
            result.update({'contents':contents,'regiDate':regiDate})
            print('키워드있음:{}'.format(keyword))
            searchList.append(keyword)
    print("searchList:",searchList,"/ searchList_TYPE:",type(searchList))

    allSentences=contents.split(".")
    findSentences=[]
    count=1
    for index,sentence in enumerate(allSentences):
        for keyword in keywords:
            if sentence.find(keyword)>=0:
                findSentences.append(str(count)+"."+sentence)
                count+=1
                break
    print("findSentences:",findSentences,"/ findSentences_TYPE:",type(findSentences))

    if len(searchList)>=1:
        searchResult=True
    else:
        searchResult=False

    # print("키워드없음")
    return searchResult,result,searchList,findSentences

def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print('Error: Creating directory. ' + directory)

def cosineSimilarityDic(data):
    # 제목 추출
    titles = [item['title'] for item in data]

    # TF-IDF Vectorizer 객체 생성 및 제목을 TF-IDF 벡터로 변환
    vectorizer = TfidfVectorizer()
    tfidf_matrix = vectorizer.fit_transform(titles)

    # 코사인 유사도 계산
    cosine_sim = cosine_similarity(tfidf_matrix)

    # 유사도가 0.4 이상인 요소 찾기
    threshold = 0.4
    unique_indices = set(range(len(data)))  # 모든 인덱스를 초기 집합으로 설정

    for i in range(len(cosine_sim)):
        for j in range(i + 1, len(cosine_sim)):
            if cosine_sim[i][j] >= threshold:
                if j in unique_indices:
                    unique_indices.remove(j)  # 중복 제거

    # 중복 제거된 데이터 리스트 생성
    unique_data = [data[i] for i in unique_indices]
    return unique_data

class Thread(QThread):
    cnt = 0
    user_signal = pyqtSignal(str)  # 사용자 정의 시그널 2 생성

    def __init__(self, parent,startDate,endDate,countLimit,fname,keywords):  # parent는 WndowClass에서 전달하는 self이다.(WidnowClass의 인스턴스)
        super().__init__(parent)
        self.parent = parent  # self.parent를 사용하여 WindowClass 위젯을 제어할 수 있다.
        self.startDate=startDate
        self.endDate=endDate
        self.countLimit=countLimit
        self.fname=fname
        self.keywords=keywords


    def run(self):

        fnameCompany = 'companyName.xlsx'

        # 키워드 가져오기
        # keywords, searchKeywords = GetKeywords(self.fname)
        keywords=[self.keywords]


        # 회사이름 가져오기
        companyNames = GetCompanyName(fnameCompany)

        # 기사목록 가져오기
        timeNow = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")


        # filename = 'RESULT_{}_{}_from_{}_to_{}.csv'.format(timeNow,searchKeywords.replace(" | ",""), self.startDate.replace(".", ""), self.endDate.replace(".", ""))
        wb=openpyxl.Workbook()

        for index,keyword in enumerate(keywords):
            searchKeywords=keyword
            # 'Sheet'라는 이름의 기본 시트가 있는지 확인하고 삭제
            if 'Sheet' in wb.sheetnames:
                std = wb['Sheet']
                wb.remove(std)

            #신규로 시트 생성
            ws=wb.create_sheet(title=searchKeywords)
            title = ['검색어', '제목검출어\n(기업명)', '본문검출어\n(키워드)', '발행일', '기사제목', '기사본문','발췌문장', '신문사', '해당링크']
            ws.append(title)

            #기사 ID값들 전부 가져오기
            resultList = GetArticles(keywords,searchKeywords, companyNames, self.startDate, self.endDate, self.countLimit,self) # 기사 아이디 가져오기
            
            print("가져오기완료")
            print("중복제거전갯수:", len(resultList))

            # url 밸류값을 기준으로 중복된 딕셔너리 제거
            resultList = remove_duplicate_dicts(resultList, key="url")
            print("중복제거완료")
            if len(resultList)==0:
                break

            #유사한 문장 합쳐주는 로직

            resultList=cosineSimilarityDic(resultList)

            print("유사문장제거완료")


            with open('resultList.json', 'w',encoding='utf-8-sig') as f:
                json.dump(resultList, f, indent=2,ensure_ascii=False)

            text = "중복제거된갯수:{}".format(len(resultList))
            self.user_signal.emit(text)
            dataList=[]
            finalList = []

            #개별 신문 기사에 대해서 상세 내용 조회 반복문
            for index, result in enumerate(resultList):
                text = "{}/{}번째 상세 페이지 확인중...".format(index + 1, len(resultList))
                print(text)
                self.user_signal.emit(text)
                isKeyword, resultData,searchList,findSentences = GetDetail(searchKeywords, result)
                if isKeyword == True:
                    finalList.append(resultData)
                    data = [searchKeywords, ",".join(resultData['companyFind']),",".join(searchList), resultData['regiDate'],resultData['title'],
                            resultData['contents'],"\n".join(findSentences), resultData['company'].replace("언론사 선정", ""), resultData['url']]
                    dataList.append(data)
                    # writer.writerow(data)
                    with open('finalList.json', 'w', encoding='utf-8-sig') as f:
                        json.dump(finalList, f, indent=2, ensure_ascii=False)
                print("===============================")
                time.sleep(random.randint(10, 20) * 0.1)

            with open('dataList.json', 'w',encoding='utf-8-sig') as f:
                json.dump(dataList, f, indent=2,ensure_ascii=False)

            # 두 번째 요소를 기준으로 오름차순으로 정렬
            dataList = sorted(dataList, key=lambda x: x[1])

            for data in dataList:
                ws.append(data)
                wb.save('RESULT_{}_from_{}_to_{}.xlsx'.format(timeNow,self.startDate.replace(".", ""), self.endDate.replace(".", "")))


            # 전체 행에 자동 줄바꿈 설정
            for index,row in enumerate(ws.iter_rows()):
                if index==0:
                    for index,cell in enumerate(row):
                        cell.alignment = Alignment(wrapText=True,vertical='center',horizontal='center')
                else:
                    for index,cell in enumerate(row):
                        if index<=4:
                            cell.alignment = Alignment(wrapText=True,vertical='center',horizontal='center')
                        elif 5<=index<=6:
                            cell.alignment = Alignment(wrapText=True,vertical='top')
                        else:
                            cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')



            # 행 높이 지정
            for i in range(2, ws.max_row + 1):
                ws.row_dimensions[i].height = 100

            # 열의 폭 설정
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 60
            ws.column_dimensions['F'].width = 50
            ws.column_dimensions['G'].width = 50
            ws.column_dimensions['H'].width = 15

            # 첫 번째 행 고정
            ws.freeze_panes = 'A2'

            # # 전체 열에 필터 적용
            ws.auto_filter.ref = ws.dimensions

            # 노란색 채우기 스타일 생성
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            # 첫 번째 행의 모든 셀에 노란색 적용
            for cell in ws[1]:  # 첫 번째 행의 모든 셀에 접근
                cell.fill = yellow_fill


            wb.save('RESULT_{}_from_{}_to_{}.xlsx'.format(timeNow, self.startDate.replace(".", ""),self.endDate.replace(".", "")))
        text = "작업완료"
        self.user_signal.emit(text)

    def stop(self):
        pass


class Example(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.path = "C:"
        self.index = None
        self.setupUi(self)
        self.setSlot()
        self.show()
        QApplication.processEvents()
        #오늘날짜지정
        today = QDate.currentDate()
        self.dateEdit_2.setDate(today)
        #일주일전지정
        today = QDate.currentDate()
        one_week_ago = today.addDays(-7)  # 현재 날짜에서 7일 전 날짜 계산
        self.dateEdit.setDate(one_week_ago)
        self.fname=""
        self.lineEdit_2.setText("10")

    def start(self):
        # self.fname="keyword.xlsx"
        self.keywords=self.lineEdit.text()
        selected_date = self.dateEdit.date()
        self.startDate = selected_date.toString("yyyy.MM.dd")
        print("self.startDate:",self.startDate,"/ self.startDate_TYPE:",type(self.startDate),len(self.startDate))
        selected_date = self.dateEdit_2.date()
        self.endDate = selected_date.toString("yyyy.MM.dd")
        print("self.endDate:",self.endDate,"/ self.endDate_TYPE:",type(self.endDate),len(self.endDate))
        self.countLimit=int(self.lineEdit_2.text())
        print("self.countLimit:",self.countLimit,"/ self.countLimit_TYPE:",type(self.countLimit))
        print("self.fname:",self.fname,"/ self.fname_TYPE:",type(self.fname),len(self.fname))
        self.x = Thread(self,self.startDate,self.endDate,self.countLimit,self.fname,self.keywords)
        self.x.user_signal.connect(self.slot1)  # 사용자 정의 시그널2 슬롯 Connect
        self.x.start()

    def slot1(self, data1):  # 사용자 정의 시그널1에 connect된 function
        self.textEdit.append(str(data1))
    def slot2(self):  # 사용자 정의 시그널1에 connect된 function
        QMessageBox.information(self, "에러창", "키워드 파일을 먼저 불러와주세요.")


    def setSlot(self):
        pass


    def find(self):
        print("find")
        self.fname = QFileDialog.getOpenFileName(self, "Open file", './')[0]
        print(self.fname)
        self.lineEdit.setText(self.fname)

    def setIndex(self, index):
        pass

    def quit(self):
        QCoreApplication.instance().quit()


app = QApplication([])
ex = Example()
sys.exit(app.exec_())


# fname1 = 'keyword.xlsx'
# fname2 = 'companyName.xlsx'
# searchStart='2023.11.01'
# searchEnd='2023.11.12'
# countLimit=100
# #키워드 가져오기
# keywords,searchKeywords=GetKeywords(fname1)
# #회사이름 가져오기
# companyNames=GetCompanyName(fname2)
# #기사목록 가져오기
# timeNow=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
# filename='SEARCH_RESULT_{}_keyword_{}_from_{}_to_{}.csv'.format(timeNow,searchKeywords.replace(" | ",""),searchStart.replace(".",""),searchEnd.replace(".",""))
# f=open(filename, 'w',encoding='utf-8-sig',newline='')
# writer=csv.writer(f)
# title=['키워드','기업명','발행일','기사본문','신문사','해당링크']
# writer.writerow(title)
#
#
# resultList=GetArticles(keywords,searchKeywords,companyNames,searchStart,searchEnd,countLimit)
# print("가져오기완료")
# print("중복제거전갯수:",len(resultList))
# newResultList=[]
# for result in resultList:
#     if result not in newResultList:
#         newResultList.append(result)
# resultList=newResultList
# print("중복제거된갯수:",len(resultList))
#
# finalList=[]
# for index,result in enumerate(resultList):
#     text="{}/{}번째 확인중...".format(index+1,len(resultList))
#     print(text)
#     isKeyword,resultData=GetDetail(keywords,result)
#     if isKeyword==True:
#         finalList.append(resultData)
#         data=[resultData['keyword'],resultData['companyName'],resultData['regiDate'],resultData['contents'],resultData['company'].replace("언론사 선정",""),resultData['url']]
#         writer.writerow(data)
#     with open('finalList.json', 'w',encoding='utf-8-sig') as f:
#         json.dump(finalList, f, indent=2,ensure_ascii=False)
#     print("===============================")
#     time.sleep(random.randint(10,20)*0.1)
# f.close()
